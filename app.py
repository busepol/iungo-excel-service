from flask import Flask, request, send_file
import io
import datetime
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from fpdf import FPDF # Pure Python PDF library (no Railway errors!)

app = Flask(__name__)

# ==========================================
# EXCEL STYLING & HELPERS
# ==========================================
DARK_BLUE  = "203764"   
MID_BLUE   = "305496"   
LIGHT_BLUE = "D9E1F2"   
YELLOW     = "FFF2CC"   
WHITE      = "FFFFFF"
GRAY       = "F2F2F2"   
BORDER_COL = "B7B7B7"

def get_border():
    s = Side(style="thin", color=BORDER_COL)
    return Border(left=s, right=s, top=s, bottom=s)

def style_cell(ws, cell, value=None, bold=False, bg=None, fg="000000",
               sz=9, ha="left", va="center", wrap=False, locked=True,
               italic=False, fmt=None):
    c = ws[cell] if isinstance(cell, str) else cell
    if value is not None:
        c.value = value
    c.font = Font(name="Arial", bold=bold, size=sz, color=fg, italic=italic)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    c.border = get_border()
    c.protection = Protection(locked=locked)
    if fmt:
        c.number_format = fmt
    return c

def build_iungo_xlsx(order):
    wb = Workbook()
    ws = wb.active
    ws.title = "ORDINE"

    # Column Widths
    widths = [6, 20, 20, 40, 8, 12, 15, 12, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Headers
    ws.merge_cells("A1:I1")
    style_cell(ws, "A1", "REGALIDEA S.R.L. — ORDINE FORNITORE",
               bold=True, sz=14, bg=DARK_BLUE, fg=WHITE, ha="center")

    ws.merge_cells("A2:I2")
    style_cell(ws, "A2", "I campi GIALLI sono da compilare dal fornitore. Non modificare la struttura.",
               italic=True, sz=9, bg=MID_BLUE, fg=WHITE, ha="center")

    # Info Block
    # Left side: rows 4-10 (7 fields: doc number, issue date, delivery date,
    #            destination, phone, email, fax)
    # Right side: rows 4-7 (4 fields: customer name, customer vat,
    #             supplier name, supplier vat) — rows 8-10 left blank on right
    info_left = [
        (4,  "N. DOCUMENTO:",   order.get("docNumber", ""),    GRAY,   True),
        (5,  "DATA EMISSIONE:", order.get("issueDate", ""),    GRAY,   True),
        (6,  "DATA CONSEGNA:",  order.get("deliveryDate", ""), YELLOW, False),
        (7,  "DESTINAZIONE:",   order.get("destination", ""),  YELLOW, False),
        (8,  "TELEFONO:",       order.get("phone", ""),        YELLOW, False),
        (9,  "EMAIL:",          order.get("email", ""),        YELLOW, False),
        (10, "FAX:",            order.get("fax", ""),          YELLOW, False),
    ]
    info_right = [
        (4, "NOME CLIENTE:",    order.get("customerName", ""), YELLOW, False),
        (5, "P.IVA CLIENTE:",   order.get("customerVat", ""),  YELLOW, False),
        (6, "NOME FORNITORE:",  "REGALIDEA S.R.L.",            GRAY,   True),
        (7, "P.IVA FORNITORE:", "IT00926410010",               GRAY,   True),
    ]

    for r, label, val, bg, lck in info_left:
        ws.merge_cells(f"A{r}:B{r}")
        style_cell(ws, f"A{r}", label, bold=True, fg=DARK_BLUE, ha="right")
        ws.merge_cells(f"C{r}:E{r}")
        style_cell(ws, f"C{r}", val, bg=bg, locked=lck)

    for r, label, val, bg, lck in info_right:
        ws.cell(row=r, column=7, value=label).font = Font(bold=True, color=DARK_BLUE)
        ws.cell(row=r, column=7).alignment = Alignment(horizontal="right")
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)
        style_cell(ws, ws.cell(row=r, column=8), val, bg=bg, locked=lck)

    # Fill right-side cells (rows 8-10) with plain gray so they look intentional
    for r in range(8, 11):
        for col in [7, 8]:
            c = ws.cell(row=r, column=col)
            c.fill = PatternFill("solid", start_color=GRAY)
            c.border = get_border()
        ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=9)

    # Totals — row 11 (one row after the last info row)
    TOTALS_ROW = 11
    style_cell(ws, f"B{TOTALS_ROW}", "TOTALE QUANTITÀ:", bold=True, fg=DARK_BLUE, ha="right")
    ws.merge_cells(f"C{TOTALS_ROW}:E{TOTALS_ROW}")
    style_cell(ws, f"C{TOTALS_ROW}", f"=SUM(F{TOTALS_ROW+4}:F500)",
               bg=LIGHT_BLUE, bold=True, ha="center", fmt="#,##0")

    style_cell(ws, f"G{TOTALS_ROW}", "TOTALE IMPORTO:", bold=True, fg=DARK_BLUE, ha="right")
    ws.merge_cells(f"H{TOTALS_ROW}:I{TOTALS_ROW}")
    style_cell(ws, f"H{TOTALS_ROW}", f"=SUM(I{TOTALS_ROW+4}:I500)",
               bg=LIGHT_BLUE, bold=True, ha="center", fmt='€ #,##0.00')

    # Table Headers — row 12 (one blank spacer row 12, headers row 13)
    HEADER_ROW = 13
    # Row 12: blank spacer — style it nicely
    for col in range(1, 10):
        c = ws.cell(row=12, column=col)
        c.fill = PatternFill("solid", start_color=GRAY)
        c.border = get_border()

    headers = ["#", "CODICE ARTICOLO", "CODICE FORNITORE", "DESCRIZIONE",
               "U.M.", "QTÀ", "PREZZO LORDO €", "SCONTO %", "IMPORTO NETTO €"]
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=HEADER_ROW, column=col, value=text)
        style_cell(ws, c, bold=True, bg=DARK_BLUE, fg=WHITE, ha="center", sz=8)

    # Line Items — start row 14
    FIRST_DATA_ROW = HEADER_ROW + 1
    items = order.get("items", [])
    for idx, item in enumerate(items):
        row = FIRST_DATA_ROW + idx
        netto_formula = f"=F{row}*G{row}*(1-H{row}/100)"
        
        row_data = [
            (1, idx+1,                       GRAY,       True),
            (2, item.get("itemCode", ""),     YELLOW,     False),
            (3, item.get("supplierCode", ""), YELLOW,     False),
            (4, item.get("description", ""),  YELLOW,     False),
            (5, item.get("um", "PZ"),         YELLOW,     False),
            (6, item.get("qty", 0),           YELLOW,     False),
            (7, item.get("grossPrice", 0),    YELLOW,     False),
            (8, item.get("discount", 0),      YELLOW,     False),
            (9, netto_formula,                LIGHT_BLUE, True),
        ]

        for col, val, bg, lck in row_data:
            fmt = '€ #,##0.00' if col in [7, 9] else ("#,##0" if col == 6 else None)
            style_cell(ws, ws.cell(row=row, column=col), val, bg=bg, locked=lck,
                       fmt=fmt, ha="center" if col != 4 else "left")

    # Empty filler rows
    for r in range(FIRST_DATA_ROW + len(items), 101):
        for c in range(1, 10):
            bg = YELLOW if 2 <= c <= 8 else (GRAY if c == 1 else LIGHT_BLUE)
            style_cell(ws, ws.cell(row=r, column=c), bg=bg, locked=not (2 <= c <= 8))
            if c == 9:
                ws.cell(row=r, column=9).value = f"=F{r}*G{r}*(1-H{r}/100)"
                ws.cell(row=r, column=9).number_format = '€ #,##0.00'

    # Protection
    ws.protection.sheet = True
    ws.protection.password = "regalidea2026"
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ==========================================
# ROUTE 1: THE EXCEL GENERATOR
# ==========================================
@app.route("/generate", methods=["POST"])
def generate():
    data = request.get_json()
    if not data: return {"error": "No JSON body"}, 400
    
    order = data.get("order", data)
    now = datetime.datetime.now()
    doc_num = f"ORD-{now.year}-{str(int(time.time()))[-6:]}"
    issue_date = now.strftime("%d/%m/%Y")

    items = []
    for i in order.get("items", []):
        item_code = i.get("itemCode") or i.get("codice") or ""
        supplier_code = i.get("supplierCode", "") 
        
        items.append({
            "itemCode": item_code,
            "supplierCode": supplier_code,
            "description": i.get("descrizione") or i.get("description", ""),
            "um": i.get("um", "PZ"),
            "qty": i.get("quantitaGabrielli") or i.get("quantita") or i.get("quantity", 0),
            "grossPrice": i.get("prezzo") or i.get("fascia", 0),
            "discount": i.get("sconto", 0),
        })

    order_data = {
        "docNumber": doc_num,
        "issueDate": issue_date,
        "deliveryDate": order.get("deliveryDate") or order.get("deliveryWindow") or "",
        "destination": order.get("luogoConsegna") or order.get("destination") or "",
        "customerName": order.get("customer") or order.get("pointOfSale") or "",
        "customerVat": order.get("customerVat", ""),
        "phone": order.get("phone", ""),
        "email": order.get("email", ""),
        "fax": order.get("fax", ""),
        "items": items
    }

    buf = build_iungo_xlsx(order_data)
    filename = f"IUNGO_ORDER_{order_data['customerName'].replace(' ', '_')}.xlsx"

    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ==========================================
# PDF STYLING & CLASS
# ==========================================
class RegalideaPDF(FPDF):
    def header(self):
        self.set_fill_color(32, 55, 100) # #203764
        self.set_text_color(255, 255, 255)
        self.set_font("helvetica", "B", 14)
        self.cell(0, 10, "REGALIDEA S.R.L. - ORDINE FORNITORE", align="C", fill=True, ln=True)
        self.set_font("helvetica", "I", 10)
        self.cell(0, 6, "I campi GIALLI sono da compilare dal fornitore. Non modificare la struttura.", align="C", fill=True, ln=True)
        self.ln(5)

# ==========================================
# ROUTE 2: THE PDF GENERATOR
# ==========================================
@app.route("/generate-pdf", methods=["POST"])
def generate_pdf():
    try:
        data = request.get_json(force=True) 
        if not data: return {"error": "No JSON body received"}, 400
        
        order = data.get("order", data)

        pdf = RegalideaPDF(orientation="L", unit="mm", format="A4")
        pdf.add_page()

        # --- COLOR PALETTE (Matching Excel) ---
        pdf.set_font("helvetica", "B", 9)
        blue_text = (32, 55, 100)
        yellow_fill = (255, 242, 204)
        gray_fill = (242, 242, 242)
        light_blue_fill = (217, 225, 242)
        
        # --- PREPARE DATA ---
        now = datetime.datetime.now()
        doc_num = str(order.get("orderNumber") or order.get("docNumber") or f"ORD-{now.year}-{str(int(time.time()))[-6:]}")
        issue_date = now.strftime("%d/%m/%Y")
        
        customer_name = str(order.get("customer") or order.get("pointOfSale") or "")
        customer_vat = str(order.get("customerVat") or order.get("piva") or "")
        del_date = str(order.get("deliveryDate") or order.get("deliveryWindow") or "")[:10]
        dest = str(order.get("luogoConsegna") or order.get("destination") or "")[:65]
        phone = str(order.get("phone", ""))
        email = str(order.get("email", ""))
        fax = str(order.get("fax", ""))

        # ==========================================
        # INFO BLOCK — two-column layout
        # Left column: label (35mm) + value (80mm)
        # Right column: label (35mm) + value (127mm)
        # Rows where the right column has no content get a gray placeholder.
        # ==========================================
        ROW_H = 6

        def left_label(text):
            pdf.set_text_color(*blue_text)
            pdf.set_font("helvetica", "B", 9)
            pdf.cell(35, ROW_H, text, border=1, align="R")

        def left_value(text, fill_color=gray_fill):
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("helvetica", "", 9)
            pdf.set_fill_color(*fill_color)
            pdf.cell(80, ROW_H, text, border=1, fill=True)

        def right_label(text):
            pdf.set_text_color(*blue_text)
            pdf.set_font("helvetica", "B", 9)
            pdf.cell(35, ROW_H, text, border=1, align="R")

        def right_value(text, fill_color=yellow_fill, newline=True):
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("helvetica", "", 9)
            pdf.set_fill_color(*fill_color)
            pdf.cell(127, ROW_H, text, border=1, fill=True, ln=newline)

        def right_empty(newline=True):
            """Blank gray cell on the right when there is no matching field."""
            pdf.set_fill_color(*gray_fill)
            pdf.cell(162, ROW_H, "", border=1, fill=True, ln=newline)

        # Row 1 — N. DOCUMENTO / NOME CLIENTE
        left_label("N. DOCUMENTO:")
        left_value(doc_num, gray_fill)
        right_label("NOME CLIENTE:")
        right_value(customer_name[:65], yellow_fill)

        # Row 2 — DATA EMISSIONE / P.IVA CLIENTE
        left_label("DATA EMISSIONE:")
        left_value(issue_date, gray_fill)
        right_label("P.IVA CLIENTE:")
        right_value(customer_vat, yellow_fill)

        # Row 3 — DATA CONSEGNA / NOME FORNITORE
        left_label("DATA CONSEGNA:")
        left_value(del_date, yellow_fill)
        right_label("NOME FORNITORE:")
        right_value("REGALIDEA S.R.L.", gray_fill)

        # Row 4 — DESTINAZIONE / P.IVA FORNITORE
        left_label("DESTINAZIONE:")
        left_value(dest, yellow_fill)
        right_label("P.IVA FORNITORE:")
        right_value("IT00926410010", gray_fill)

        # Row 5 — TELEFONO / (no right field)
        left_label("TELEFONO:")
        left_value(phone, yellow_fill)
        right_empty()

        # Row 6 — EMAIL / (no right field)
        left_label("EMAIL:")
        left_value(email, yellow_fill)
        right_empty()

        # Row 7 — FAX / (no right field)
        left_label("FAX:")
        left_value(fax, yellow_fill)
        right_empty()

        pdf.ln(5)

        # ==========================================
        # PRE-CALCULATE TOTALS 
        # ==========================================
        items = order.get("items", [])
        total_qty = 0
        total_amount = 0
        
        for i in items:
            try: qty = float(i.get("quantitaGabrielli") or i.get("quantita") or i.get("quantity") or 0)
            except ValueError: qty = 0.0
                
            try: price = float(i.get("prezzo") or i.get("fascia") or i.get("grossPrice") or 0)
            except ValueError: price = 0.0
                
            try: discount = float(i.get("sconto") or i.get("discount") or 0)
            except ValueError: discount = 0.0
                
            total_qty += qty
            total_amount += (qty * price * (1 - (discount / 100)))

        # ==========================================
        # TOTALS ROW
        # ==========================================
        pdf.set_font("helvetica", "B", 9)
        pdf.set_text_color(*blue_text)
        
        pdf.cell(35, 8, "TOTALE QUANTITÀ:", border=0, align="R")
        pdf.set_text_color(0, 0, 0)
        pdf.set_fill_color(*light_blue_fill)
        pdf.cell(80, 8, str(int(total_qty)), border=1, align="C", fill=True)

        pdf.set_text_color(*blue_text)
        pdf.cell(35, 8, "TOTALE IMPORTO:", border=0, align="R")
        pdf.set_text_color(0, 0, 0)
        pdf.cell(127, 8, f"EUR {total_amount:.2f}", border=1, align="C", fill=True)
        pdf.ln(10)

        # ==========================================
        # TABLE HEADERS
        # ==========================================
        pdf.set_fill_color(32, 55, 100)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("helvetica", "B", 8)
        
        cols = [("#", 10), ("COD. ARTICOLO", 25), ("COD. FORNITORE", 27), 
                ("DESCRIZIONE", 98), ("U.M.", 10), ("QTÀ", 15), 
                ("PREZZO EUR", 22), ("SCONTO %", 20), ("NETTO EUR", 50)]

        for col_name, width in cols:
            pdf.cell(width, 7, col_name, border=1, align="C", fill=True)
        pdf.ln()

        # ==========================================
        # TABLE ROWS
        # ==========================================
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("helvetica", "", 8)

        for idx, i in enumerate(items, 1):
            code = str(i.get("itemCode") or i.get("codice") or "")
            sup_code = str(i.get("supplierCode") or "")
            desc = str(i.get("descrizione") or i.get("description", ""))[:70]
            
            try: qty = float(i.get("quantitaGabrielli") or i.get("quantita") or i.get("quantity") or 0)
            except ValueError: qty = 0.0
                
            try: price = float(i.get("prezzo") or i.get("fascia") or i.get("grossPrice") or 0)
            except ValueError: price = 0.0
            
            try: discount = float(i.get("sconto") or i.get("discount") or 0)
            except ValueError: discount = 0.0

            net = qty * price * (1 - (discount / 100))

            # Index Col
            pdf.set_fill_color(*gray_fill)
            pdf.cell(10, 6, str(idx), border=1, align="C", fill=True)

            # Main Cols
            pdf.set_fill_color(*yellow_fill)
            pdf.cell(25, 6, code, border=1, align="C", fill=True)
            pdf.cell(27, 6, sup_code, border=1, align="C", fill=True)
            pdf.cell(98, 6, desc, border=1, fill=True)
            pdf.cell(10, 6, str(i.get("um", "PZ")), border=1, align="C", fill=True)
            pdf.cell(15, 6, str(int(qty)), border=1, align="C", fill=True)
            pdf.cell(22, 6, f"EUR {price:.2f}", border=1, align="R", fill=True)
            pdf.cell(20, 6, str(int(discount)), border=1, align="C", fill=True)
            
            # Netto Col
            pdf.set_fill_color(*light_blue_fill)
            pdf.cell(50, 6, f"EUR {net:.2f}", border=1, align="R", fill=True)
            pdf.ln()

        pdf_buffer = io.BytesIO()
        pdf.output(pdf_buffer)
        pdf_buffer.seek(0)
        
        safe_name = customer_name.replace(' ', '_').replace('/', '')
        if not safe_name: safe_name = "Sconosciuto"
        
        return send_file(
            pdf_buffer,
            mimetype="application/pdf",
            as_attachment=True,
            download_name=f"IUNGO_ORDER_{safe_name}.pdf"
        )
        
    except Exception as e:
        import traceback
        return {"CRASH_REPORT": str(e), "DETAILS": traceback.format_exc()}, 500

if __name__ == "__main__":
    import os
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
